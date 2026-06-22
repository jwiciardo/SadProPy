from pathlib import Path
import warnings
from openpyxl import load_workbook

class InputReader:
    def __init__(self, inputfile_path: str | Path):
        self.inputfile_path = Path(inputfile_path)
        if not self.inputfile_path.exists():
            raise FileNotFoundError(f"File not found: {self.inputfile_path}")
        with warnings.catch_warnings():
            warnings.filterwarnings("ignore", message="Data Validation extension is not supported*")
            self.workbook = load_workbook(inputfile_path, data_only=True)
    
    # CENTRAL FUNCTION: READER
    def read_inputfile(self, sheet_name: str, start_row: int = 1):
        if sheet_name not in self.workbook.sheetnames:
            raise ValueError(f"Sheet '{sheet_name}' not found. Available: {self.workbook.sheetnames}")
        worksheet = self.workbook[sheet_name]
        rows = list(worksheet.values)
        headers = rows[start_row - 1]
        records = []
        for row in rows[start_row:]:
            record = dict(zip(headers, row))
            if all(v is None for v in record.values()): # Skip completely empty rows
                continue
            records.append(record)
        return records