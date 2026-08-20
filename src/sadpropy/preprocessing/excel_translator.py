import warnings
import numpy as np
from openpyxl import load_workbook
from .preprocessing_dictionary import (
    material_type_dict,
    material_model_dict,
    material_definition_dict,
    section_shape_dict,
    section_model_dict,
    section_definition_dict,
    integration_type_dict,
    element_type_dict,
    loadcase_type_dict,
)
from .preprocessing_class_index import SectionModel
from .preprocessing_dataclass import (
    FilePathInformation,
    ProjectInformation,
    AnalysisPreferences,
    Materials,
    FrameSections,
    SlabSections,
    Storeys,
)
from .preprocessing_object._shell import _generate_shell_connectivity
from .preprocessing_object._section import _compute_section_properties, _trace_aggregated_sections
from .preprocessing_object._load import _get_concentrated_elemental_loads, _get_distributed_elemental_loads, _get_shell_to_elemental_loads
from ..utility import ConverterToInternalUnits, UserDefinedUnits
from ..utility import TagManager
from ..utility.exception import ValidationError
from ..utility.filepath import get_filepath

class ExcelReader:
    def __init__(self, inputfile_path):
        if not inputfile_path.exists():
            raise FileNotFoundError(f"File not found: {inputfile_path}")
        with warnings.catch_warnings():
            warnings.filterwarnings("ignore", message="Data Validation extension is not supported*")
            self._workbook = load_workbook(inputfile_path, data_only=True)
    
    # MAIN METHOD: EXCELREADER
    def read(self, sheet_name="", orientation="records", start_row=1):
        if sheet_name not in self._workbook.sheetnames:
            raise ValidationError(f"Sheet '{sheet_name}' not found in Excel file")
        worksheet = self._workbook[sheet_name]
        rows = list(worksheet.values)
        headers = list(rows[start_row - 1])
        while headers and headers[-1] is None:
            headers.pop()
        if orientation == "records":
            data = []
        elif orientation == "columns":
            data = {header: [] for header in headers}
        else:
            raise ValidationError("Orientation must be either 'records' or 'columns'")

        nrows = 0
        for row in rows[start_row:]:
            row = row[:len(headers)]
            if all(v is None for v in row): # Skip completely empty rows
                continue
            nrows += 1
            if orientation == "records":
                data.append(dict(zip(headers, row)))
            elif orientation == "columns":
                for header, value in zip(headers, row):
                    data[header].append(value)
        return data, nrows
    
    def read_preferences_sheet(self, sheet_name="", start_row=1, required_preferences=None):
        data, nrows = self.read(sheet_name=sheet_name, orientation="records", start_row=start_row)
        values = {row["Item"]: row["Value"] for row in data}

        if required_preferences is not None:
            for preference in required_preferences:
                value = values.get(preference)
                if value is None or str(value).strip() == "":
                    raise ValidationError(f"'{preference}' in sheet '{sheet_name}' cannot be empty")
        return values

class ExcelTranslator:
    def __init__(self, inputfile_path):
        # FILE PATH
        self._inputfile_path = inputfile_path
        self._parent_path, self._output_path, self._logfile_path = get_filepath(inputfile_path)

        # CORE
        self._reader = ExcelReader(inputfile_path=self._inputfile_path)
        self._units = self._translate_userdefined_units()
        self._to_internalunits = ConverterToInternalUnits(units=self._units)
        self._tagmanager = TagManager()
        
    # MAIN METHOD: EXCEL TRANSLATOR
    def translate(self):
        filepath_information = self._translate_filepath_information()
        project_information = self._translate_project_information()
        analysis_preferences = self._translate_analysis_preferences()
        materials = self._translate_materials()
        frame_sections = self._translate_frame_sections(materials=materials)
        slab_sections = self._translate_slab_sections(materials=materials)
        node_objects, storeys = self._translate_node_objects(project_information=project_information)
        element_objects = self._translate_element_objects(node_objects=node_objects, sections=frame_sections)
        shell_objects = self._translate_shell_objects(element_objects=element_objects, slab_sections=slab_sections)
        restraints = self._translate_restraints(node_objects=node_objects)
        nodal_loads = self._translate_nodal_loads()
        concentrated_elemental_loads = self._translate_concentrated_elemental_loads()
        distributed_elemental_loads = self._translate_distributed_elemental_loads(element_objects=element_objects)
        shell_to_elemental_loads = self._translate_shell_to_elemental_loads(element_objects=element_objects, shell_objects=shell_objects)
        return {
            "Filepath Information": filepath_information,
            "Project Information": project_information,
            "Userdefined Units": self._units,
            "Analysis Preferences": analysis_preferences,
            "Materials": materials,
            "Frame Sections": frame_sections,
            "Slab Sections": slab_sections,
            "Storeys": storeys,
            "Node Objects": node_objects,
            "Element Objects": element_objects,
            "Shell Objects": shell_objects,
            "Restraints": restraints,
            "Nodal Loads": nodal_loads,
            "Concentrated Elemental Loads": concentrated_elemental_loads,
            "Distributed Elemental Loads": distributed_elemental_loads,
            "Shell to Elemental Loads": shell_to_elemental_loads,
        }

    # HELPER METHOD
    def _validate_data(self, nrows, sheet_name, mandatory=True):
        if nrows > 0: # Set condition if number of rows in data > 0 then return True
            return True
        if mandatory: # Set condition if mandatory is True then raise validation error (This condition will be run when number of rows in data = 0)
            raise ValidationError(f"Sheet '{sheet_name}' is mandatory but contains no data.")
        return False

    def _validate_duplicate_value(self, col_data, col_name=""):
        unique_names, counts = np.unique(col_data, return_counts=True)
        duplicates = unique_names[counts > 1]
        if duplicates.size > 0:
            raise ValidationError(f"Duplicate {col_name}(s): {', '.join(duplicates)}")
    
    def _generate_storeys(self, storey_elevations):
        storey_name = [] # Predefined storeys list
        height = [] # Predefined storeys height
        elevation = [] # Predefined storeys elevation
        for idx, elev in reversed(list(enumerate(storey_elevations))): # Loop over storey elevations
            if idx == 0: # Set condition if index == 0
                storey_name.append("Base")# If True define storey name as "Base" and height = 0.0
                height.append(np.float64(0.0))
            else:
                storey_name.append(f"Storey{idx}") # If False define storey name as "Storey{index}" and storey height
                height.append(elev - storey_elevations[idx - 1])
            elevation.append(elev)
        storey_name = np.asarray(storey_name, dtype="U15")
        height = np.asarray(height, dtype=np.float64)
        elevation = np.asarray(elevation, dtype=np.float64)
        storeys = Storeys(
            name=storey_name,
            height=height,
            elevation=elevation,
        ) # Store storeys data as dataclass
        return storeys

    def _modify_empty_values(self, values, converter=None, dtype=object, filled_values=np.nan):
        modified_values = np.where(
            [value is not None for value in values],
            converter(values=values) if converter is not None else np.asarray(values, dtype=dtype), 
            filled_values,
        )
        return modified_values

    def _group_typical_columns(self, columns, converter=None, dtype=object):
        group_values = np.column_stack([
            np.asarray(column, dtype=dtype) if converter is None else converter(values=column)
            for column in columns
        ])
        return group_values

    # SUPPORTING METHODS
    def _translate_filepath_information(self):
        filepath_information = FilePathInformation(
            parent_path = self._parent_path,
            output_path = self._output_path,
            inputfile_path = self._inputfile_path,
            logfile_path = self._logfile_path
        ) # Storing filepath information to dataclass
        return filepath_information

    def _translate_project_information(self):
        values = self._reader.read_preferences_sheet(
            sheet_name="Project Information",
            start_row=6,
            required_preferences=["Model Dimensional Space"],
        ) # Reading Sheet "Project Information" in the Input file
        project_information = ProjectInformation(
            name = str(values["Project Name"]),
            desc = str(values["Project Description"]),
            ndim = int(3 if str(values["Model Dimensional Space"]) == "3D-Space" else 2),
        ) # Storing project information to dataclass
        return project_information
    
    def _translate_userdefined_units(self):
        values = self._reader.read_preferences_sheet(
            sheet_name="User Defined Units",
            start_row=9,
            required_preferences=[
                "Force",
                "Length",
                "Mass",
                "Stress",
                "Time",
                "Angle",
            ],
        ) # Reading Sheet "User Defined Units" in the Input file
        userdefined_units = UserDefinedUnits(
            force = str(values["Force"]),
            length = str(values["Length"]),
            mass = str(values["Mass"]),
            stress = str(values["Stress"]),
            time = str(values["Time"]),
            angle = str(values["Angle"]),
        ) # Storing units to dataclass
        return userdefined_units

    def _translate_analysis_preferences(self):
        values = self._reader.read_preferences_sheet(
            sheet_name="Analysis Preferences",
            start_row=6,
            required_preferences=[
                "Nonlinear Analysis",
                "P-Delta",
                "LL Mass Factor",
            ],
        ) # Reading Sheet "Analysis Preferences" in the Input file
        analysis_preferences = AnalysisPreferences(
            is_nonlinear_analysis = str(values["Nonlinear Analysis"]).strip().lower() == "yes",
            is_pdelta = str(values["P-Delta"]).strip().lower() == "yes",
            liveload_mass_factor = float(values["LL Mass Factor"]),
        ) # Storing analysis preferences to dataclass
        return analysis_preferences

    def _translate_materials(self):
        sheet_name = "Materials"
        data, n = self._reader.read(
            sheet_name=sheet_name, 
            orientation="columns", 
            start_row=32,
        ) # Reading Sheet "Materials" in the Input file
        self._validate_data(nrows=n, sheet_name=sheet_name, mandatory=True)
        index = np.arange(n, dtype=np.int32)
        mat_name = np.asarray(data["Material Name"], dtype="U32")
        self._validate_duplicate_value(col_data=mat_name, col_name="Material Name")
        mat_tag = np.asarray(self._tagmanager.add(category="Material", n=n, names=mat_name), dtype=np.int32)
        mat_type = np.asarray([material_type_dict[value.strip().title()] for value in data["Material Type"]], dtype=np.int8)
        mat_model = np.asarray([material_model_dict[value.strip()] for value in data["Material Model"]], dtype=np.int8)

        # Translate material properties
        mat_def = [material_definition_dict[(mtype, model)] for mtype, model in zip(mat_type, mat_model)]
        max_columns = max(definition.properties.Count for definition in mat_def)
        properties = np.full((n, max_columns), np.nan, dtype=np.float64)
        unique_definitions = list(dict.fromkeys(mat_def))
        for definition in unique_definitions:
            mask = np.asarray([d is definition for d in mat_def], dtype=bool)
            selected_data = {key: np.asarray(value)[mask] for key, value in data.items()}
            translated_props = definition.translate(data=selected_data, converter=self._to_internalunits)
            properties[mask, :definition.properties.Count] = translated_props
        materials = Materials(
            index = index,
            mat_name = mat_name,
            mat_tag = mat_tag,
            mat_type = mat_type,
            mat_model = mat_model,
            mat_def = mat_def,
            properties = properties,
        ) # Storing materials data to dataclass
        return materials
    
    def _translate_frame_sections(self, materials):
        sheet_name = "Frame Sections"
        data, n = self._reader.read(
            sheet_name=sheet_name, 
            orientation="columns", 
            start_row=19,
        ) # Reading Sheet "Frame Sections" in the Input file
        self._validate_data(nrows=n, sheet_name=sheet_name, mandatory=True)
        index = np.arange(n, dtype=np.int32)
        sec_name = np.asarray(data["Section Name"], dtype="U32")
        self._validate_duplicate_value(col_data=sec_name, col_name="Section Name")
        sec_tag = np.asarray(self._tagmanager.add(category="Section", n=n, names=sec_name), dtype=np.int32)
        sec_shape = np.asarray([section_shape_dict[value.strip().title()] for value in data["Section Shape"]], dtype=np.int8)
        sec_model = np.asarray([section_model_dict[value.strip().title()] for value in data["Section Model"]], dtype=np.int8)
        mat_columns = ["Material", "Material2", "Material3", "Material4", "Material5", "Material6"]
        mats_idx = np.column_stack([
            materials.name_to_idx(names=data[column]) for column in mat_columns
        ])
        mat_type = np.full(n, -1, dtype=np.int8)
        for i in range(n):
            mat_type[i] = materials.mat_type[mats_idx[i][np.argmax(mats_idx[i] != -1)]]
        integration_type = np.asarray([
            integration_type_dict[value.strip().title()]
            if value is not None and value.strip() else -1
            for value in data["Integration Type"]],
            dtype=np.int8
        )
        integration_points = np.asarray([
            value if value is not None else -1
            for value in data["Integration Points"]],
            dtype=np.int32
        )
        integration_mask = integration_type != -1 # Integration masking
        integration_tag = np.full(n, -1, dtype=np.int32)
        if np.any(integration_mask):
            integration_tag[integration_mask] = np.asarray(self._tagmanager.add(category="Beam Integration", n=np.count_nonzero(integration_mask), names=sec_name[integration_mask]), dtype=np.int32)  
        section_lookup = dict(zip(sec_name, index))
        aggregated_sec_idx = np.asarray([section_lookup[name]
            if name is not None and str(name).strip() else -1
            for name in data["Aggregated Section"]],
            dtype=np.int32
        )

        # Translate section dimensions
        aggregator_sec_mask = sec_model == SectionModel.Aggregator # Aggregator section masking
        normal_sec_mask = ~aggregator_sec_mask # Normal section masking
        resolved_sec_idx = _trace_aggregated_sections(
            aggregated_sec_idx=aggregated_sec_idx,
            aggregator_mask=aggregator_sec_mask,
            sec_name=sec_name,
        )

        sec_def = np.empty(n, dtype=object)
        normal_sec_idx = np.flatnonzero(normal_sec_mask)
        for i in normal_sec_idx:
            sec_def[i] = section_definition_dict[
                (mat_type[i], sec_shape[i], sec_model[i])
            ]
        sec_def[aggregator_sec_mask] = sec_def[resolved_sec_idx[aggregator_sec_mask]]
        max_columns = max(definition.dimensions.Count for definition in sec_def[normal_sec_mask])
        dimensions = np.full((n, max_columns), np.nan, dtype=np.float64)
        unique_definitions = list(dict.fromkeys(sec_def[normal_sec_mask]))
        for definition in unique_definitions:
            mask = (
                normal_sec_mask 
                & np.fromiter(
                    (value is definition for value in sec_def),
                    dtype=bool,
                    count=n,
                ))
            selected_data = {key: np.asarray(value)[mask] for key, value in data.items()}
            translated_dims = definition.translate(data=selected_data, converter=self._to_internalunits)
            dimensions[mask, :definition.dimensions.Count] = translated_dims
        dimensions[aggregator_sec_mask] = dimensions[resolved_sec_idx[aggregator_sec_mask]]

        # Compute section properties
        AMod = np.asarray(data["AMod"], dtype=np.float64)
        AvyMod = np.asarray(data["AvyMod"], dtype=np.float64)
        AvzMod = np.asarray(data["AvzMod"], dtype=np.float64)
        IzMod = np.asarray(data["IzMod"], dtype=np.float64)
        IyMod = np.asarray(data["IyMod"], dtype=np.float64)
        JxxMod = np.asarray(data["JxxMod"], dtype=np.float64)
        normal_sec_props = _compute_section_properties(
            section_definitions=sec_def[normal_sec_mask],
            dimensions=dimensions[normal_sec_mask],
        )
        properties = np.full((n, 12), np.nan, dtype=np.float64)
        properties[normal_sec_mask] = np.column_stack(normal_sec_props)
        properties[:, 0] *= AMod
        properties[:, 1] *= AvyMod
        properties[:, 2] *= AvzMod
        properties[:, 3] *= IzMod
        properties[:, 4] *= IyMod
        properties[:, 5] *= JxxMod
        properties[aggregator_sec_mask] = properties[resolved_sec_idx[aggregator_sec_mask]]
        frame_sections = FrameSections(
            index = index,
            sec_name = sec_name,
            sec_tag = sec_tag,
            sec_shape = sec_shape,
            sec_model = sec_model,
            sec_def = sec_def,
            mats_idx = mats_idx,
            mat_type = mat_type,
            integration_type = integration_type,
            integration_points = integration_points,
            integration_tag = integration_tag,
            aggregated_sec_idx=aggregated_sec_idx,
            dimensions=dimensions,
            properties = properties,
        ) # Storing sections data to dataclass
        return frame_sections

    def _translate_slab_sections(self, materials):
        sheet_name = "Slab Sections"
        data, n = self._reader.read(
            sheet_name=sheet_name, 
            orientation="columns", 
            start_row=6,
        ) # Reading Sheet "Slab Sections" in the Input file
        if not self._validate_data(nrows=n, sheet_name=sheet_name, mandatory=False):
            slab_sections = SlabSections.empty()
            return slab_sections
        index = np.arange(n, dtype=np.int32)
        sec_name = np.asarray(data["Section Name"], dtype="U32")
        self._validate_duplicate_value(col_data=sec_name, col_name="Section Name")
        mat_columns = ["Material"]
        mats_idx = np.column_stack([
            materials.name_to_idx(names=data[column]) for column in mat_columns
        ])
        mat_type = np.empty(n, dtype=np.int8)
        for i in range(n):
            if mats_idx[i, 0] >= 0:
                j = 0
            else:
                continue
            mat_type[i] = materials.mat_type[mats_idx[i, j]]
        t = self._to_internalunits.length(values=data["t"])
        dimensions = np.column_stack((
            t,
        ))
        slab_sections = SlabSections(
            index = index,
            sec_name = sec_name,
            mats_idx = mats_idx,
            mat_type=mat_type,
            dimensions = dimensions,
        ) # Storing slab sections data to dataclass
        return slab_sections
    
    def _translate_node_objects(self, project_information):
        sheet_name = "Node Objects"
        data, n = self._reader.read(
            sheet_name=sheet_name, 
            orientation="columns", 
            start_row=7,
        ) # Reading Sheet "Node Objects" in the Input file
        self._validate_data(nrows=n, sheet_name=sheet_name, mandatory=True)
        index = np.arange(n, dtype=np.int32)
        unique_name = np.asarray(data["Unique Name"], dtype="U15")
        self._validate_duplicate_value(col_data=unique_name, col_name="Unique Name")
        coord_columns = [data["X"], data["Y"], data["Z"]]
        coords = self._group_typical_columns(columns=coord_columns, converter=self._to_internalunits.length, dtype=np.float64)
        name_to_idx = dict(zip(unique_name, index))
        node_objects = {
            "Index": index,
            "Unique Name": unique_name,
            "Coordinates": coords,
            "Name to Index": name_to_idx,
        } # Storing node object data to dictionary
        ndim = project_information.ndim # Retrieve number of dimensional space
        storey_elevations = np.unique(coords[:, 2]) if ndim == 3 else np.unique(coords[:, 1]) # Determine storey elevations
        storeys = self._generate_storeys(storey_elevations=storey_elevations) # Generating Storey data from storey elevations
        return node_objects, storeys

    def _translate_element_objects(self, node_objects, sections):
        sheet_name = "Element Objects"
        data, n = self._reader.read(
            sheet_name=sheet_name, 
            orientation="columns", 
            start_row=15,
        ) # Reading Sheet "Element Objects" in the Input file
        self._validate_data(nrows=n, sheet_name=sheet_name, mandatory=True)
        index = np.arange(n, dtype=np.int32)
        unique_name = np.asarray(data["Unique Name"], dtype="U15")
        self._validate_duplicate_value(col_data=unique_name, col_name="Unique Name")
        iend_node = np.asarray(data["I-End"], dtype="U15")
        jend_node = np.asarray(data["J-End"], dtype="U15")
        node_name_to_idx = node_objects["Name to Index"]
        end_nodes_idx = np.column_stack((
            np.fromiter((node_name_to_idx[name] for name in iend_node), dtype=np.int32, count=n),
            np.fromiter((node_name_to_idx[name] for name in jend_node), dtype=np.int32, count=n),
        ))
        element_type = np.asarray([element_type_dict[value.strip().title()] for value in data["Element Type"]], dtype=np.int8)
        sec_idx = sections.name_to_idx(names=data["Section"])
        iend_coords = node_objects["Coordinates"][end_nodes_idx[:, 0]]
        jend_coords = node_objects["Coordinates"][end_nodes_idx[:, 1]]
        length = np.linalg.norm(jend_coords - iend_coords, axis=1)
        is_auto_end_offsets = np.fromiter((
            str(x).strip().lower() == "auto from connectivity"
            for x in data["End Offset"]),
            dtype=bool,
            count=n,
        )
        rigid_zone_factor = np.asarray(data["Rigid Zone Factor"], dtype=np.float64)
        offset_length_columns = ["I-End Offset Length", "J-End Offset Length"]
        offsets_length = np.column_stack([
            self._modify_empty_values(values=data[column], converter=self._to_internalunits.length, dtype=np.float64, filled_values=0.0) for column in offset_length_columns
        ])
        is_zero_length_element = np.fromiter((
            str(x).strip().lower() == "yes"
            for x in data["Zero Length Element"]),
            dtype=bool,
            count=n,
        )
        name_to_idx = dict(zip(unique_name, index))
        element_objects = {
            "Index": index,
            "Unique Name": unique_name,
            "End Nodes Index": end_nodes_idx,
            "Element Type": element_type,
            "Section Index": sec_idx,
            "Length": length,
            "Is Auto End Offsets": is_auto_end_offsets,
            "Rigid Zone Factor": rigid_zone_factor,
            "Offsets Length": offsets_length,
            "Is Zero Length Element": is_zero_length_element,
            "Name to Index": name_to_idx,
        } # Storing element objects data to dictionary
        return element_objects

    def _translate_shell_objects(self, element_objects, slab_sections):
        sheet_name = "Shell Objects"
        data, n = self._reader.read(
            sheet_name=sheet_name, 
            orientation="columns", 
            start_row=10,
        ) # Reading Sheet "Shell Objects" in the Input file
        self._validate_data(nrows=n, sheet_name=sheet_name, mandatory=True)
        index = np.arange(n, dtype=np.int32)
        unique_name = np.asarray(data["Unique Name"], dtype="U15")
        self._validate_duplicate_value(col_data=unique_name, col_name="Unique Name")
        edge_columns = [data["Edge 1"], data["Edge 2"], data["Edge 3"], data["Edge 4"]]
        edges_name = self._group_typical_columns(columns=edge_columns, dtype="U15")
        edges_idx = np.empty((n, 4), dtype=np.int32)
        vertices_idx = np.empty((n, 4), dtype=np.int32)
        for i in range(n):
            edges_idx[i], vertices_idx[i] = _generate_shell_connectivity(
                edges_name=edges_name[i],
                element_objects=element_objects,
                shell_name=unique_name[i],
            )
        element_type = np.asarray([element_type_dict[value.strip().title()] for value in data["Element Type"]], dtype=np.int8)
        sec_idx = slab_sections.name_to_idx(names=data["Section"])
        name_to_idx = dict(zip(unique_name, index))
        shell_objects = {
            "Index": index,
            "Unique Name": unique_name,
            "Edges Index": edges_idx,
            "Vertices Index": vertices_idx,
            "Element Type": element_type,
            "Section Index": sec_idx,
            "Name to Index": name_to_idx,
        } # Storing shell objects data to dictionary
        return shell_objects
    
    def _translate_restraints(self, node_objects):
        sheet_name="Restraints"
        data, n = self._reader.read(
            sheet_name=sheet_name, 
            orientation="columns", 
            start_row=10,
        ) # Reading Sheet "Restraints" in the Input file
        self._validate_data(nrows=n, sheet_name=sheet_name, mandatory=True)
        node_name_to_idx = node_objects["Name to Index"]
        node_name = np.asarray(data["Node"], dtype="U15")
        node_idx = np.fromiter((node_name_to_idx[name] for name in node_name), dtype=np.int32, count=n)
        dof_columns = [data["UX"], data["UY"], data["UZ"], data["RX"], data["RY"], data["RZ"]]
        dofs = self._group_typical_columns(columns=dof_columns, dtype=np.int32)
        restraints = {
            "Node Index": node_idx,
            "DOFs": dofs,
        } # Storing restraints data to dictionary
        return restraints

    def _translate_nodal_loads(self):
        sheet_name="Nodal Loads"
        data, n = self._reader.read(
            sheet_name=sheet_name, 
            orientation="columns", 
            start_row=11,
        ) # Reading Sheet "Nodal Loads" in the Input file
        if not self._validate_data(nrows=n, sheet_name=sheet_name, mandatory=False):
            nodal_loads = []
            return nodal_loads
        node_name = np.asarray(data["Node"], dtype="U15")
        loadcase_type = np.asarray([loadcase_type_dict[value.strip().title()] for value in data["Load Case"]], dtype=np.int8)
        force_columns = ["FX", "FY", "FZ"]
        moment_columns = ["MX", "MY", "MZ"]
        loads = np.column_stack(
            [self._modify_empty_values(values=data[column], converter=self._to_internalunits.force_nodal_load, dtype=np.float64, filled_values=0.0) for column in force_columns] +
            [self._modify_empty_values(values=data[column], converter=self._to_internalunits.moment_nodal_load, dtype=np.float64, filled_values=0.0) for column in moment_columns]
        )
        nodal_loads = {
            "Node Name": node_name,
            "Load Case": loadcase_type,
            "Loads": loads,
        } # Storing Nodal Loads data to dictionary
        return nodal_loads

    def _translate_concentrated_elemental_loads(self):
        sheet_name="Concentrated Elemental Loads"
        data, n = self._reader.read(
            sheet_name=sheet_name, 
            orientation="columns", 
            start_row=14,
        ) # Reading Sheet "Concentrated Elemental Loads" in the Input file
        if not self._validate_data(nrows=n, sheet_name=sheet_name, mandatory=False):
            concentrated_elemental_loads = []
            return concentrated_elemental_loads
        element_name = np.asarray(data["Element"], dtype="U15")
        loadcase_type = np.asarray([loadcase_type_dict[value.strip().title()] for value in data["Load Case"]], dtype=np.int8)
        load_direction = np.asarray(data["Direction"], dtype="U15")
        load_columns = ["Load 1", "Load 2", "Load 3", "Load 4"]
        loads = np.column_stack([
            self._modify_empty_values(values=data[column], converter=self._to_internalunits.concentrated_elemental_load, dtype=np.float64, filled_values=np.nan) for column in load_columns
        ])
        location_columns = ["Location 1", "Location 2", "Location 3", "Location 4"]
        locations = np.column_stack([
            self._modify_empty_values(values=data[column], converter=self._to_internalunits.length, dtype=np.float64, filled_values=np.nan) for column in location_columns
        ])
        modified_element_name, modified_loadcase_type, modified_load_direction, modified_location, modified_load = _get_concentrated_elemental_loads(
            element_name=element_name,
            loadcase_type=loadcase_type,
            load_direction=load_direction,
            locations=locations,
            loads=loads,
        )
        concentrated_element_loads = {
            "Element Name": modified_element_name,
            "Load Case": modified_loadcase_type,
            "Direction": modified_load_direction,
            "Location": modified_location,
            "Load": modified_load,
        } # Storing Concentrated Element Loads data to dictionary
        return concentrated_element_loads

    def _translate_distributed_elemental_loads(self, element_objects):
        sheet_name="Distributed Elemental Loads"
        data, n = self._reader.read(
            sheet_name=sheet_name, 
            orientation="columns", 
            start_row=15,
        ) # Reading Sheet "Distributed Elemental Loads" in the Input file
        if not self._validate_data(nrows=n, sheet_name=sheet_name, mandatory=False):
            distributed_elemental_loads = []
            return distributed_elemental_loads
        element_name = np.asarray(data["Element"], dtype="U15")
        loadcase_type = np.asarray([loadcase_type_dict[value.strip().title()] for value in data["Load Case"]], dtype=np.int8)
        load_direction = np.asarray(data["Direction"], dtype="U15")
        uniform_load = self._modify_empty_values(values=data["Uniform Load"], converter=self._to_internalunits.distributed_elemental_load, dtype=np.float64, filled_values=np.nan)
        load_columns = ["Load 1", "Load 2", "Load 3", "Load 4"]
        loads = np.column_stack([
            self._modify_empty_values(values=data[column], converter=self._to_internalunits.distributed_elemental_load, dtype=np.float64, filled_values=np.nan) for column in load_columns
        ])
        location_columns = ["Location 1", "Location 2", "Location 3", "Location 4"]
        locations = np.column_stack([
            self._modify_empty_values(values=data[column], converter=self._to_internalunits.length, dtype=np.float64, filled_values=np.nan) for column in location_columns
        ])
        modified_element_name, modified_loadcase_type, modified_load_direction, modified_location, modified_load = _get_distributed_elemental_loads(
            element_name=element_name,
            element_objects=element_objects,
            loadcase_type=loadcase_type,
            load_direction=load_direction,
            locations=locations,
            uniform_load=uniform_load,
            loads=loads,
        )
        distributed_elemental_loads = {
            "Element Name": modified_element_name,
            "Load Case": modified_loadcase_type,
            "Direction": modified_load_direction,
            "Location": modified_location,
            "Load": modified_load,
        } # Storing Distributed Element Loads data to dictionary
        return distributed_elemental_loads

    def _translate_shell_to_elemental_loads(self, element_objects, shell_objects):
        sheet_name="Shell to Elemental Loads"
        data, n = self._reader.read(
            sheet_name=sheet_name, 
            orientation="columns", 
            start_row=7,
        ) # Reading Sheet "Shell to Elemental Loads" in the Input file
        if not self._validate_data(nrows=n, sheet_name=sheet_name, mandatory=False):
            shell_to_elemental_loads = []
            return shell_to_elemental_loads
        shell_name = np.asarray(data["Shell"], dtype="U15")
        loadcase_type = np.asarray([loadcase_type_dict[value.strip().title()] for value in data["Load Case"]], dtype=np.int8)
        load_direction = np.asarray(data["Direction"], dtype="U15")
        load = self._modify_empty_values(values=data["Load"], converter=self._to_internalunits.shell_load, dtype=np.float64, filled_values=np.nan)
        modified_shell_name, modified_edge_name, modified_loadcase_type, modified_load_direction, modified_location, modified_load = _get_shell_to_elemental_loads(
            shell_name=shell_name,
            element_objects=element_objects,
            shell_objects=shell_objects,
            loadcase_type=loadcase_type,
            load_direction=load_direction,
            load=load,
        )
        shell_to_elemental_loads = {
            "Shell Name": modified_shell_name,
            "Edge Name": modified_edge_name,
            "Load Case": modified_loadcase_type,
            "Direction": modified_load_direction,
            "Location": modified_location,
            "Load": modified_load,
        } # Storing Shell to Elemental Loads data to dictionary
        return shell_to_elemental_loads