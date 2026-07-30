import warnings
import numpy as np
from openpyxl import load_workbook
from ._preproc_dataclass import (
    FilePathInformation,
    ProjectInformation,
    AnalysisPreferences,
    Materials,
    Mat_Concrete04,
    Mat_Steel02,
    Mat_MinMax,
    Mat_IMK,
    FrameSections,
    Sec_Fiber,
    Sec_Aggregator,
    SlabSections,
    Storeys,
)
from ._preproc_class import (
    MaterialProperties,
    Concrete04Properties,
    Steel02Properties,
    MinMaxProperties,
    IMKProperties,
    FrameSectionProperties,
    FiberSectionProperties,
    SectionAggregatorProperties,
    SlabSectionProperties,
)
from ._materialdata import get_material_properties
from ._sectiondata import get_section_properties
from sadpropy.utility import (
    ConverterToInternalUnits,
    UserDefinedUnits,
    section_properties,
    fibersection_properties,
)
from sadpropy.utility._exceptions import ValidationError
from sadpropy.utility._filepath import get_filepath
from sadpropy.utility.helperfunc import get_edges_and_vertices_from_surface


class ExcelReader:
    def __init__(self, inputfile_path):
        if not inputfile_path.exists():
            raise FileNotFoundError(f"File not found: {inputfile_path}")
        with warnings.catch_warnings():
            warnings.filterwarnings("ignore", message="Data Validation extension is not supported*")
            self._workbook = load_workbook(inputfile_path, data_only=True)
    
    # MAIN METHOD: EXCELREADER
    def read(self, sheet_name="", start_row=1):
        if sheet_name not in self._workbook.sheetnames:
            raise ValidationError(f"Sheet '{sheet_name}' not found in Excel file")
        worksheet = self._workbook[sheet_name]
        rows = list(worksheet.values)
        headers = list(rows[start_row - 1])
        while headers and headers[-1] is None:
            headers.pop()
        data = []
        for row in rows[start_row:]:
            row = row[:len(headers)]
            record = dict(zip(headers, row))
            if all(v is None for v in record.values()): # Skip completely empty rows
                continue
            data.append(record)
        return data
    
    def read_preferences_sheet(self, sheet_name="", start_row=1, required_preferences=None):
        data = self.read(sheet_name=sheet_name, start_row=start_row)
        values = {row["Item"]: row["Value"] for row in data}

        if required_preferences is not None:
            for preference in required_preferences:
                value = values.get(preference)
                if value is None or str(value).strip() == "":
                    raise ValidationError(f"'{preference}' in sheet '{sheet_name}' cannot be empty.")
        return values

class ExcelTranslator:
    def __init__(self):
        # FILE PATH
        self._parent_path, self._output_path, self._inputfile_path, self._logfile_path = get_filepath()

        # CORE
        self._reader = ExcelReader(inputfile_path=self._inputfile_path)
        self._units = self._translate_userdefined_units()
        self._to_internalunits = ConverterToInternalUnits(units=self._units)
        
        # DICTIONARY LIST
        self._mats_list = []
        self._secs_list = []

    # MAIN METHOD: EXCEL TRANSLATOR
    def translate(self):
        filepath_information = self._translate_filepath_information()
        project_information = self._translate_project_information()
        analysis_preferences = self._translate_analysis_preferences()
        materials = self._translate_materials()
        mat_concrete04 = self._translate_mat_concrete04()
        mat_steel02 = self._translate_mat_steel02()
        mat_minmax = self._translate_mat_minmax()
        mat_imk = self._translate_mat_imk()
        frame_sections = self._translate_frame_sections()
        sec_fiber = self._translate_sec_fiber()
        sec_aggregator = self._translate_sec_aggregator()
        slab_sections = self._translate_slab_sections()
        point_objects, storeys = self._translate_point_objects(project_information=project_information)
        line_objects = self._translate_line_objects(point_objects=point_objects)
        surface_objects = self._translate_surface_objects(
            line_objects=line_objects,
            slab_sections=slab_sections,
        )
        restraints = self._translate_restraints(point_objects=point_objects)
        return {
            "Filepath Information": filepath_information,
            "Project Information": project_information,
            "Userdefined Units": self._units,
            "Analysis Preferences": analysis_preferences,
            "Materials": materials,
            "Mat: Concrete04": mat_concrete04,
            "Mat: Steel02": mat_steel02,
            "Mat: Minmax": mat_minmax,
            "Mat: IMK": mat_imk,
            "Materials List": self._mats_list,
            "Frame Sections": frame_sections,
            "Sec: Fiber": sec_fiber,
            "Sec: Aggregator": sec_aggregator,
            "Sections List": self._secs_list,
            "Slab Sections": slab_sections,
            "Storeys": storeys,
            "Point Objects": point_objects,
            "Line Objects": line_objects,
            "Surface Objects": surface_objects,
            "Restraints": restraints,
        }

    # HELPER METHOD
    def _validate_data(self, data, sheet_name, mandatory=True):
        if len(data) > 0: # Set condition if number of rows in data > 0 then return True
            return True
        if mandatory: # Set condition if mandatory is True then raise validation error (This condition will be run when number of rows in data = 0)
            raise ValidationError(f"Sheet '{sheet_name}' is mandatory but contains no data.")
        return False
    
    def _generate_storeys(self, storey_elevations):
        storeys = {} # Predefined storeys dictionary
        for idx, elev in reversed(list(enumerate(storey_elevations))): # Loop over storey elevations
            if idx == 0: # Set condition if index == 0
                storey_name = "Base" # If True define storey name as "Base" and height = 0.0
                height = np.float64(0.0) 
            else:
                storey_name = f"Storey{idx}" # If False define storey name as "Storey{index}" and storey height
                height = elev - storey_elevations[idx - 1]
            storeys[storey_name] = Storeys(
                name=storey_name,
                height=height,
                elevation=elev,
            ) # Store storeys data as dataclass
        return storeys

    def _retrieve_material_index(self, mats_name):
        if isinstance(mats_name, str): # Set condition if materials name is a string return list of materials name
            mats_name = [mats_name]
        mat_class = np.empty(len(mats_name), dtype=np.int32) # Predefined material class array
        mat_idx = np.empty(len(mats_name), dtype=np.int32) # Predefined material index array 
        mat = [] # Predefined materials dataclass list
        for i, mat_name in enumerate(mats_name): # Loop over materials name
            found = False # Predefined found material in materials list
            for cls, material in enumerate(self._mats_list): # Loop over materials list
                idx = material.name_to_idx.get(mat_name) # Retrieve material index
                if idx is not None: # Set condition if material index is not None
                    mat_class[i] = cls # Return material class, material index, and materials dataclass for index i
                    mat_idx[i] = idx
                    mat.append(material)
                    found = True # Set found to be True
                    break
            if not found: # Set condition if found is False return validation error
                raise ValidationError(f"Material '{mat_name}' not found")
        if len(mats_name) == 1: # Set condition if number of rows in materials name is 1
            return mat_class[0], mat[0], mat_idx[0] # Return material class, material index, and materials dataclass for first index
        return mat_class, mat, mat_idx

    def _retrieve_section_index(self, secs_name):
        if isinstance(secs_name, str): # Set condition if sections name is a string return list of sections name
            secs_name = [secs_name]
        sec_class = np.empty(len(secs_name), dtype=np.int32) # Predefined section class array
        sec_idx = np.empty(len(secs_name), dtype=np.int32) # Predefined section index array 
        sec = [] # Predefined sections dataclass list
        for i, sec_name in enumerate(secs_name): # Loop over sections name
            found = False # Predefined found section in sections list
            for cls, section in enumerate(self._secs_list): # Loop over sections list
                idx = section.name_to_idx.get(sec_name) # Retrieve section index
                if idx is not None: # Set condition if section index is not None
                    sec_class[i] = cls # Return section class, section index, and sections dataclass for index i
                    sec_idx[i] = idx
                    sec.append(section)
                    found = True # Set found to be True
                    break
            if not found: # Set condition if found is False return validation error
                raise ValidationError(f"Section '{sec_name}' not found")
        if len(secs_name) == 1: # Set condition if number of rows in sections name is 1
            return sec_class[0], sec[0], sec_idx[0] # Return section class, section index, and sections dataclass for first index
        return sec_class, sec, sec_idx
    
    def _retrieve_slabsection_index(self, secs_name, secs_list): # Retrieve Slab section 
        if isinstance(secs_name, str): # Set condition if sections name is a string return list of sections name
            secs_name = [secs_name]
        sec_class = np.empty(len(secs_name), dtype=np.int32) # Predefined section class array
        sec_idx = np.empty(len(secs_name), dtype=np.int32) # Predefined section index array 
        sec = [] # Predefined sections dataclass list
        for i, sec_name in enumerate(secs_name): # Loop over sections name
            found = False # Predefined found section in sections list
            for cls, section in enumerate(secs_list): # Loop over sections list
                idx = section.name_to_idx.get(sec_name) # Retrieve section index
                if idx is not None: # Set condition if section index is not None
                    sec_class[i] = cls # Return section class, section index, and sections dataclass for index i
                    sec_idx[i] = idx
                    sec.append(section)
                    found = True # Set found to be True
                    break
            if not found: # Set condition if found is False return validation error
                raise ValidationError(f"Section '{sec_name}' not found")
        if len(secs_name) == 1: # Set condition if number of rows in sections name is 1
            return sec_class[0], sec[0], sec_idx[0] # Return section class, section index, and sections dataclass for first index
        return sec_class, sec, sec_idx
        
    # SUPPORTING METHODS
    def _translate_filepath_information(self):
        filepath_information = FilePathInformation(
            parent_path = self._parent_path,
            output_path = self._output_path,
            inputfile_path = self._inputfile_path,
            logfile_path = self._logfile_path
        ) # Storing filepath information to dataclass
        return filepath_information

    def _translate_project_information(self):
        values = self._reader.read_preferences_sheet(
            sheet_name="Project Information",
            start_row=6,
            required_preferences=["Model Dimensional Space"],
        ) # Reading Sheet "Project Information" in the Input file
        project_information = ProjectInformation(
            name = str(values["Project Name"]),
            desc = str(values["Project Description"]),
            ndim = int(3 if str(values["Model Dimensional Space"]) == "3-Dimensional" else 2),
        ) # Storing project information to dataclass
        return project_information
    
    def _translate_userdefined_units(self):
        values = self._reader.read_preferences_sheet(
            sheet_name="User Defined Units",
            start_row=9,
            required_preferences=[
                "Force",
                "Length",
                "Mass",
                "Stress",
                "Time",
                "Angle",
            ],
        ) # Reading Sheet "User Defined Units" in the Input file
        userdefined_units = UserDefinedUnits(
            force = str(values["Force"]),
            length = str(values["Length"]),
            mass = str(values["Mass"]),
            stress = str(values["Stress"]),
            time = str(values["Time"]),
            angle = str(values["Angle"]),
        ) # Storing units to dataclass
        return userdefined_units

    def _translate_analysis_preferences(self):
        values = self._reader.read_preferences_sheet(
            sheet_name="Analysis Preferences",
            start_row=6,
            required_preferences=[
                "Nonlinear Analysis",
                "P-Delta",
                "LL Mass Factor",
            ],
        ) # Reading Sheet "Analysis Preferences" in the Input file
        analysis_preferences = AnalysisPreferences(
            is_nonlinear_analysis = str(values["Nonlinear Analysis"]),
            is_pdelta = str(values["P-Delta"]),
            liveload_mass_factor = float(values["LL Mass Factor"]),
        ) # Storing analysis preferences to dataclass
        return analysis_preferences

    def _translate_materials(self):
        sheet_name = "Materials"
        data = self._reader.read(sheet_name=sheet_name, start_row=13) # Reading Sheet "Materials" in the Input file
        self._validate_data(data=data, sheet_name=sheet_name, mandatory=True)
        n = len(data)
        index = np.arange(n, dtype=np.int32)
        mat_name = np.empty(n, dtype="U32")
        mat_type = np.empty(n, dtype="U15")
        properties = np.zeros((n, len(MaterialProperties)), dtype=np.float64)
        name_to_idx = {}
        for i, row in enumerate(data):
            name = str(row["Material Name"])
            if name in name_to_idx:
                raise ValidationError(f"Duplicate Material name '{name}'")
            name_to_idx[name] = index[i]
            mat_name[i] = name
            mattype = str(row["Material Type"])
            mat_type[i] = mattype
            Unitweight = self._to_internalunits.unitweight(value=row["Unitweight"])
            E = self._to_internalunits.stress(row["E"])
            nu = row["nu"]
            fc = self._to_internalunits.stress(value=row["fc"]) if mattype == "Concrete" else 0.0
            fy = self._to_internalunits.stress(value=row["fy"]) if mattype in ("Rebar", "Steel") else 0.0
            fu = self._to_internalunits.stress(value=row["fu"]) if mattype in ("Rebar", "Steel") else 0.0
            properties[i] = (Unitweight, E, nu, 0.0, fc, fy, fu,) # Look at MaterialProperties class in _propertiesclass.py to find definition of variables
        E, nu = properties[:, MaterialProperties.E], properties[:, MaterialProperties.nu]
        properties[:, MaterialProperties.G] = E / (2 * (1 + nu))
        materials = Materials(
            index = index,
            mat_name = mat_name,
            mat_type = mat_type,
            properties = properties,
            name_to_idx = name_to_idx,
        ) # Storing materials data to dataclass
        self._mats_list.append(materials) # Append Materials Properties into Material Lists
        return materials
    
    def _translate_mat_concrete04(self):
        sheet_name = "Mat_Concrete04"
        data = self._reader.read(sheet_name=sheet_name, start_row=12) # Reading Sheet "Mat_Concrete04" in the Input file
        if not self._validate_data(data=data, sheet_name=sheet_name, mandatory=False):
            mat_concrete04 = []
            self._mats_list.append(mat_concrete04)
            return mat_concrete04
        n = len(data)
        index = np.arange(n, dtype=np.int32)
        mat_name = np.empty(n, dtype="U32")
        mat_type = np.empty(n, dtype="U15")
        base_mat_class = np.empty(n, dtype=np.int32)
        base_mat_idx = np.empty(n, dtype=np.int32)
        mat_model = np.empty(n, dtype="U15")
        properties = np.zeros((n, len(Concrete04Properties)), dtype=np.float64)
        name_to_idx = {}
        for i, row in enumerate(data):
            name = str(row["Material Name"])
            if name in name_to_idx:
                raise ValidationError(f"Duplicate Material name '{name}'")
            name_to_idx[name] = index[i]
            mat_name[i] = name
            mat_class, _, mat_idx = self._retrieve_material_index(mats_name=str(row["Base Material"]))
            base_mat_class[i] = mat_class
            base_mat_idx[i] = mat_idx
            mat_type[i] = self._mats_list[mat_class].mat_type[mat_idx]
            mat_model[i] = str(row["Material Model"])
            fc = self._to_internalunits.stress(value=row["fc"])
            epsc, epscu = row["epsc"], row["epscu"]
            fct = self._to_internalunits.stress(value=row["fct"])
            et = row["et"] if row["et"] != 0.0 else fct * epsc / fc
            beta = row["beta"]
            properties[i] = (0.0, 0.0, 0.0, 0.0, -fc, -epsc, -epscu, fct, et, beta,) # Look at Concrete04Properties class in _propertiesclass.py to find definition of variables
        base_mat_props = get_material_properties(
            mats_list=self._mats_list,
            mat_class=base_mat_class,
            mat_idx=base_mat_idx,
            props_name=["Unitweight", "E", "nu", "G"],
        )
        properties[:, Concrete04Properties.Unitweight:Concrete04Properties.G+1] = base_mat_props
        mat_concrete04 = Mat_Concrete04(
            index = index,
            mat_name = mat_name,
            mat_type = mat_type,
            base_mat_class = base_mat_class,
            base_mat_idx = base_mat_idx,
            mat_model = mat_model,
            properties = properties,
            name_to_idx = name_to_idx,
        ) # Storing materials data to dataclass
        self._mats_list.append(mat_concrete04) # Append Mat_Concrete04 Properties into Material Lists
        return mat_concrete04

    def _translate_mat_steel02(self):
        sheet_name = "Mat_Steel02"
        data = self._reader.read(sheet_name=sheet_name, start_row=18) # Reading Sheet "Mat_Steel02" in the Input file
        if not self._validate_data(data=data, sheet_name=sheet_name, mandatory=False):
            mat_steel02 = []
            self._mats_list.append(mat_steel02)
            return mat_steel02
        n = len(data)
        index = np.arange(n, dtype=np.int32)
        mat_name = np.empty(n, dtype="U32")
        mat_type = np.empty(n, dtype="U15")
        base_mat_class = np.empty(n, dtype=np.int32)
        base_mat_idx = np.empty(n, dtype=np.int32)
        mat_model = np.empty(n, dtype="U15")
        properties = np.zeros((n, len(Steel02Properties)), dtype=np.float64)
        name_to_idx = {}
        fy = np.zeros(n, dtype=np.float64)
        b = np.zeros(n, dtype=np.float64)
        fu = np.zeros(n, dtype=np.float64)
        eu= np.zeros(n, dtype=np.float64)
        for i, row in enumerate(data):
            name = str(row["Material Name"])
            if name in name_to_idx:
                raise ValidationError(f"Duplicate Material name '{name}'")
            name_to_idx[name] = index[i]
            mat_name[i] = name
            mat_class, _, mat_idx = self._retrieve_material_index(mats_name=str(row["Base Material"]))
            base_mat_class[i] = mat_class
            base_mat_idx[i] = mat_idx
            mat_type[i] = self._mats_list[mat_class].mat_type[mat_idx]
            mat_model[i] = str(row["Material Model"])
            fy[i] = self._to_internalunits.stress(value=row["fy"])
            b[i] = row["b"]
            fu[i] = self._to_internalunits.stress(value=row["fu"])
            eu[i] = row["eu"]
            R0 = row["R0"]
            cR1 = row["cR1"]
            cR2 = row["cR2"]
            a1 = row["a1"]
            a2 = row["a2"]
            a3 = row["a3"]
            a4 = row["a4"]
            f_init = self._to_internalunits.stress(value=row["f_init"])
            properties[i] = (0.0, 0.0, 0.0, 0.0, 0.0, 0.0, R0, cR1, cR2, a1, a2, a3, a4, f_init,) # Look at Steel02Properties class in _propertiesclass.py to find definition of variables
        base_mat_props = get_material_properties(
            mats_list=self._mats_list,
            mat_class=base_mat_class,
            mat_idx=base_mat_idx,
            props_name=["Unitweight", "E", "nu", "G"],
        )
        properties[:, Steel02Properties.Unitweight:Steel02Properties.G+1] = base_mat_props
        E = properties[:, Steel02Properties.E]
        if b == 0.0:
            ey = fy / E
            eoffset = ey + 0.002
            Epy = (fu - fy)/(eu - eoffset)
            b = Epy / E
        else:
            return b
        properties[:, Steel02Properties.fy] = fy
        properties[:, Steel02Properties.b] = b
        mat_steel02 = Mat_Steel02(
            index = index,
            mat_name = mat_name,
            mat_type = mat_type,
            base_mat_class = base_mat_class,
            base_mat_idx = base_mat_idx,
            mat_model = mat_model,
            properties = properties,
            name_to_idx = name_to_idx,
        ) # Storing materials data to dataclass
        self._mats_list.append(mat_steel02) # Append Mat_Steel02 Properties into Material Lists
        return mat_steel02

    def _translate_mat_minmax(self):
        sheet_name = "Mat_MinMax"
        data = self._reader.read(sheet_name=sheet_name, start_row=8) # Reading Sheet "Mat_MinMax" in the Input file
        if not self._validate_data(data=data, sheet_name=sheet_name, mandatory=False):
            mat_minmax = []
            self._mats_list.append(mat_minmax)
            return mat_minmax
        n = len(data)
        index = np.arange(n, dtype=np.int32)
        mat_name = np.empty(n, dtype="U32")
        mat_type = np.empty(n, dtype="U15")
        base_nl_mat_class = np.empty(n, dtype=np.int32)
        base_nl_mat_idx = np.empty(n, dtype=np.int32)
        mat_model = np.empty(n, dtype="U15")
        properties = np.zeros((n, len(MinMaxProperties)), dtype=np.float64)
        name_to_idx = {}
        for i, row in enumerate(data):
            name = str(row["Material Name"])
            if name in name_to_idx:
                raise ValidationError(f"Duplicate Material name '{name}'")
            name_to_idx[name] = index[i]
            mat_name[i] = name
            mat_class, _, mat_idx = self._retrieve_material_index(mats_name=str(row["Base NL Material"]))
            base_nl_mat_class[i] = mat_class
            base_nl_mat_idx[i] = mat_idx
            mat_type[i] = self._mats_list[mat_class].mat_type[mat_idx]
            mat_model[i] = str(row["Material Model"])
            ec_max = row["ecmax"]
            et_max = row["etmax"]
            properties[i] = (0.0, 0.0, 0.0, 0.0, ec_max, et_max,) # Look at MinMaxProperties class in _propertiesclass.py to find definition of variables
        base_nl_mat_props = get_material_properties(
            mats_list=self._mats_list,
            mat_class=base_nl_mat_class,
            mat_idx=base_nl_mat_idx,
            props_name=["Unitweight", "E", "nu", "G"],
        )
        properties[:, MinMaxProperties.Unitweight:MinMaxProperties.G+1] = base_nl_mat_props
        mat_minmax = Mat_MinMax(
            index = index,
            mat_name = mat_name,
            mat_type = mat_type,
            base_nl_mat_class = base_nl_mat_class,
            base_nl_mat_idx = base_nl_mat_idx,
            mat_model = mat_model,
            properties = properties,
            name_to_idx = name_to_idx
        ) # Storing materials data to dataclass
        self._mats_list.append(mat_minmax) # Append Mat_MinMax Properties into Material Lists
        return mat_minmax
    
    def _translate_mat_imk(self):
        sheet_name = "Mat_IMK"
        data = self._reader.read(sheet_name=sheet_name, start_row=19) # Reading Sheet "Mat_IMK" in the Input file
        if not self._validate_data(data=data, sheet_name=sheet_name, mandatory=False):
            mat_imk = []
            return mat_imk
        n = len(data)
        index = np.arange(n, dtype=np.int32)
        mat_name = np.empty(n, dtype="U32")
        mat_model = np.empty(n, dtype="U15")
        properties = np.zeros((n, len(IMKProperties)), dtype=np.float64)
        name_to_idx = {}
        mu_pos = np.zeros(n, dtype=np.float64)
        mu_neg = np.zeros(n, dtype=np.float64)
        for i, row in enumerate(data):
            name = str(row["Material Name"])
            if name in name_to_idx:
                raise ValidationError(f"Duplicate Material name '{name}'")
            name_to_idx[name] = index[i]
            mat_name[i] = name
            mat_model[i] = str(row["Material Model"])
            K0 = self._to_internalunits.rotational_stiffness(value=row["K0"])
            my_pos = self._to_internalunits.moment(value=row["My_Pos"])
            my_neg = self._to_internalunits.moment(value=row["My_Neg"])
            mu_pos[i] = self._to_internalunits.moment(value=row["Mu_Pos"])
            mu_neg[i] = self._to_internalunits.moment(value=row["Mu_Neg"])
            fpr_pos, fpr_neg, a_pinch, nfactor = row["Fpr_Pos"], row["Fpr_Neg"], row["A_pinch"], row["nFactor"]
            lamda_s, lamda_c, lamda_a, lamda_k, c_s, c_c, c_a, c_k = row["Lamda_S"], row["Lamda_C"], row["Lamda_A"], row["Lamda_K"], row["c_S"], row["c_C"], row["c_A"], row["c_K"]
            theta_p_pos, theta_p_neg = self._to_internalunits.angle(value=row["theta_p_Pos"]), self._to_internalunits.angle(value=row["theta_p_Neg"])
            theta_pc_pos, theta_pc_neg, res_pos, res_neg = self._to_internalunits.angle(value=row["theta_pc_Pos"]), self._to_internalunits.angle(value=row["theta_pc_Neg"]), row["Res_Pos"], row["Res_Neg"]
            theta_u_pos, theta_u_neg, d_pos, d_neg = self._to_internalunits.angle(value=row["theta_u_Pos"]), self._to_internalunits.angle(value=row["theta_u_Neg"]), row["D_Pos"], row["D_Neg"]
            properties[i] = (
                K0, 0.0, 0.0, my_pos, my_neg, fpr_pos, fpr_neg, a_pinch, nfactor, lamda_s, lamda_c, lamda_a, lamda_k, c_s, c_c, c_a, c_k,
                theta_p_pos, theta_p_neg, theta_pc_pos, theta_pc_neg, res_pos, res_neg, theta_u_pos, theta_u_neg, d_pos, d_neg,
            ) # Look at IMKProperties class in _propertiesclass.py to find definition of variables
        K0, my_pos, my_neg, theta_p_pos, theta_p_neg = properties[:, IMKProperties.K0], properties[:, IMKProperties.My_pos], properties[:, IMKProperties.My_neg], properties[:, IMKProperties.theta_p_pos], properties[:, IMKProperties.theta_p_neg] # Recal K0, my_pos, my_neg, theta_p_pos and theta_p_neg arrays
        theta_e_pos = my_pos / K0
        theta_e_neg = my_neg / K0
        Kpy_pos = (mu_pos - my_pos) / (theta_p_pos - theta_e_pos)
        Kpy_neg = (mu_neg - my_neg) / (theta_p_neg - theta_e_neg)
        properties[:, IMKProperties.as_pos] = K0 / Kpy_pos
        properties[:, IMKProperties.as_neg] = K0 / Kpy_neg
        mat_imk = Mat_IMK(
            index = index,
            mat_name = mat_name,
            mat_model = mat_model,
            properties = properties,
            name_to_idx = name_to_idx,
        ) # Storing materials data to dataclass
        return mat_imk
    
    def _translate_frame_sections(self):
        sheet_name = "Frame Sections"
        data = self._reader.read(sheet_name=sheet_name, start_row=15) # Reading Sheet "Frame Sections" in the Input file
        self._validate_data(data=data, sheet_name=sheet_name, mandatory=True)
        n = len(data)
        index = np.arange(n, dtype=np.int32)
        sec_name = np.empty(n, dtype="U32")
        sec_shape = np.empty(n, dtype="U32")
        mats_class = np.empty((n, 1), dtype=np.int32)
        mats_idx = np.empty((n, 1), dtype=np.int32)
        mat_type = np.empty(n, dtype="U15")
        sec_model = np.empty(n, dtype="U15")
        properties = np.zeros((n, len(FrameSectionProperties)), dtype=np.float64)
        name_to_idx = {}
        AMod = np.zeros(n, dtype=np.float64)
        AvyMod = np.zeros(n, dtype=np.float64)
        AvzMod = np.zeros(n, dtype=np.float64)
        IzMod = np.zeros(n, dtype=np.float64)
        IyMod = np.zeros(n, dtype=np.float64)
        JxxMod = np.zeros(n, dtype=np.float64)
        for i, row in enumerate(data):
            name = str(row["Section Name"])
            if name in name_to_idx:
                raise ValidationError(f"Duplicate Section name '{name}'")
            name_to_idx[name] = index[i]
            sec_name[i] = name
            sec_shape[i] = str(row["Section Shape"])
            mat_class, _, mat_idx = self._retrieve_material_index(mats_name=str(row["Material"]))
            mats_class[i] = mat_class
            mats_idx[i] = mat_idx
            mat_type[i] = self._mats_list[mat_class].mat_type[mat_idx]
            sec_model[i] = str(row["Section Model"])
            h = self._to_internalunits.length(value=row["h"])
            b = self._to_internalunits.length(value=row["b"])
            AMod[i] = row["AMod"]
            AvyMod[i] = row["AvyMod"]
            AvzMod[i] = row["AvzMod"]
            IzMod[i] = row["IzMod"]
            IyMod[i] = row["IyMod"]
            JxxMod[i] = row["JxxMod"]
            properties[i] = (h, b, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0)
        (A, Avy, Avz, Iz, Iy, Jxx, alphaY, alphaZ,) = section_properties(
            sec_shape=sec_shape,
            mat_type=mat_type, 
            properties=properties,
        )
        properties[:, FrameSectionProperties.A] = AMod * A
        properties[:, FrameSectionProperties.Avy] = AvyMod * Avy
        properties[:, FrameSectionProperties.Avz] = AvzMod * Avz
        properties[:, FrameSectionProperties.Iz] = IzMod * Iz
        properties[:, FrameSectionProperties.Iy] = IyMod * Iy
        properties[:, FrameSectionProperties.Jxx] = JxxMod * Jxx
        properties[:, FrameSectionProperties.AlphaY] = alphaY
        properties[:, FrameSectionProperties.AlphaZ] = alphaZ
        frame_sections = FrameSections(
            index = index,
            sec_name = sec_name,
            sec_shape = sec_shape,
            mats_class = mats_class,
            mats_idx = mats_idx,
            mat_type = mat_type,
            sec_model = sec_model,
            properties = properties,
            name_to_idx = name_to_idx
        ) # Storing sections data to dataclass
        self._secs_list.append(frame_sections) # Append FrameSections Properties into Section Lists
        return frame_sections
    
    def _translate_sec_fiber(self):
        sheet_name = "Sec_Fiber"
        data = self._reader.read(sheet_name=sheet_name, start_row=20) # Reading Sheet "Sec_Fiber" in the Input file
        if not self._validate_data(data=data, sheet_name=sheet_name, mandatory=False):
            sec_fiber = []
            self._secs_list.append(sec_fiber)
            return sec_fiber
        n = len(data)
        index = np.arange(n, dtype=np.int32)
        sec_name = np.empty(n, dtype="U32")
        sec_shape = np.empty(n, dtype="U32")
        base_sec_class = np.empty(n, dtype=np.int32)
        base_sec_idx = np.empty(n, dtype=np.int32)
        integration_type = np.empty(n, dtype="U15")
        mats_class = np.full((n, 3), -1, dtype=np.int32)
        mats_idx = np.full((n, 3), -1, dtype=np.int32)
        mat_type = np.empty(n, dtype="U15")
        sec_model = np.empty(n, dtype="U15")
        properties = np.zeros((n, len(FiberSectionProperties)), dtype=np.float64)
        name_to_idx = {}
        for i, row in enumerate(data):
            name = str(row["Section Name"])
            if name in name_to_idx:
                raise ValidationError(f"Duplicate Section name '{name}'")
            name_to_idx[name] = index[i]
            sec_name[i] = name
            sec_class, _, sec_idx = self._retrieve_section_index(secs_name=str(row["Base Section"]))
            base_sec_class[i] = sec_class
            base_sec_idx[i] = sec_idx
            sec_shape[i] = self._secs_list[sec_class].sec_shape[sec_idx]
            integration_type[i] = str(row["Integration Type"])
            material_columns = ["Material", "Material 2", "Material 3"]
            for j, column in enumerate(material_columns):
                if row[column] is None:
                    continue
                mat_class, _, mat_idx = self._retrieve_material_index(mats_name=str(row[column]))
                mats_class[i, j] = mat_class
                mats_idx[i, j] = mat_idx
            for j in range(6):
                if mats_class[i, j] >= 0:
                    mat_type[i] = self._mats_list[mats_class[i, j]].mat_type[mats_idx[i, j]]
                    break
            sec_model[i] = str(row["Section Model"])
            cover, nbars_top, nbars_bot, nbars_int = self._to_internalunits.length(value=row["cover"]), row["nBarsTop"], row["nBarsBot"], row["nBarsInt"]
            bar_dia_hoop, bar_dia_top = self._to_internalunits.length(value=row["barDiaHoop"]), self._to_internalunits.length(row["barDiaTop"])
            bar_dia_bot, bar_dia_int = self._to_internalunits.length(value=row["barDiaBot"]), self._to_internalunits.length(row["barDiaInt"])
            properties[i] = (0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, cover, nbars_top, 
                             nbars_bot, nbars_int, bar_dia_hoop, bar_dia_top, bar_dia_bot, bar_dia_int, 0.0, 0.0, 0.0)
        base_sec_props = get_section_properties(
            secs_list=self._secs_list,
            sec_class=base_sec_class,
            sec_idx=base_sec_idx,
            props_name=["h", "b"],
        )
        properties[:, FiberSectionProperties.h:FiberSectionProperties.b+1] = base_sec_props
        (A, Avy, Avz, Iz, Iy, Jxx, Abar_top, Abar_bot, Abar_int,) = fibersection_properties(
            sec_shape=sec_shape,
            mat_type=mat_type,
            properties=properties,
        )
        properties[:, FiberSectionProperties.A] = A
        properties[:, FiberSectionProperties.Avy] = Avy
        properties[:, FiberSectionProperties.Avz] = Avz
        properties[:, FiberSectionProperties.Iz] = Iz
        properties[:, FiberSectionProperties.Iy] = Iy
        properties[:, FiberSectionProperties.Jxx] = Jxx
        properties[:, FiberSectionProperties.Abar_top] = Abar_top
        properties[:, FiberSectionProperties.Abar_bot] = Abar_bot
        properties[:, FiberSectionProperties.Abar_int] = Abar_int
        sec_fiber = Sec_Fiber(
            index = index,
            sec_name = sec_name,
            sec_shape = sec_shape,
            base_sec_class = base_sec_class,
            base_sec_idx = base_sec_idx,
            integration_type = integration_type,
            mats_class = mats_class,
            mats_idx = mats_idx,
            mat_type = mat_type,
            sec_model = sec_model,
            properties = properties,
            name_to_idx = name_to_idx,
        ) # Storing sections data to dataclass
        self._secs_list.append(sec_fiber) # Append Sec_Fiber Properties into Section Lists
        return sec_fiber
    
    def _translate_sec_aggregator(self):
        sheet_name = "Sec_Aggregator"
        data = self._reader.read(sheet_name=sheet_name, start_row=15) # Reading Sheet "Sec_Aggregator" in the Input file
        if not self._validate_data(data=data, sheet_name=sheet_name, mandatory=False):
            sec_aggregator = []
            self._secs_list.append(sec_aggregator)
            return sec_aggregator
        n = len(data)
        index = np.arange(n, dtype=np.int32)
        sec_name = np.empty(n, dtype="U32")
        sec_shape = np.empty(n, dtype="U32")
        base_sec_class = np.empty(n, dtype=np.int32)
        base_sec_idx = np.empty(n, dtype=np.int32)
        mats_class = np.full((n, 6), -1, dtype=np.int32)
        mats_idx   = np.full((n, 6), -1, dtype=np.int32)
        mat_type = np.empty(n, dtype="U15")
        sec_model = np.empty(n, dtype="U15")
        aggregated_sec_class = np.full(n, -1, dtype=np.int32)
        aggregated_sec_idx = np.full(n, -1, dtype=np.int32)
        properties = np.zeros((n, len(SectionAggregatorProperties)), dtype=np.float64)
        name_to_idx = {}
        for i, row in enumerate(data):
            name = str(row["Section Name"])
            if name in name_to_idx:
                raise ValidationError(f"Duplicate Section name '{name}'")
            name_to_idx[name] = index[i]
            sec_name[i] = name
            sec_class, _, sec_idx = self._retrieve_section_index(secs_name=str(row["Base Section"]))
            base_sec_class[i] = sec_class
            base_sec_idx[i] = sec_idx
            sec_shape[i] = self._secs_list[sec_class].sec_shape[sec_idx]
            material_columns = ["Material", "Material 2", "Material 3", "Material 4", "Material 5", "Material 6"]
            for j, column in enumerate(material_columns):
                if row[column] is None:
                    continue
                mat_class, _, mat_idx = self._retrieve_material_index(mats_name=str(row[column]))
                mats_class[i, j] = mat_class
                mats_idx[i, j] = mat_idx
            for j in range(6):
                if mats_class[i, j] >= 0:
                    mat_type[i] = self._mats_list[mats_class[i, j]].mat_type[mats_idx[i, j]]
                    break
            sec_model[i] = str(row["Section Model"])
            if row["Aggregated Section"] is not None:
                agg_sec_class, _, agg_sec_idx = self._retrieve_section_index(sec_name=str(row["Aggregated Section"]))
                aggregated_sec_class[i] = agg_sec_class
                aggregated_sec_idx[i] = agg_sec_idx
        sec_class = np.where(aggregated_sec_class >= 0, aggregated_sec_class, base_sec_class)
        sec_idx = np.where(aggregated_sec_idx >= 0, aggregated_sec_idx, base_sec_idx,)
        sec_props = get_section_properties(
            secs_list=self._secs_list,
            sec_class=sec_class,
            sec_idx=sec_idx,
            props_name=["h", "b", "A", "Avy", "Avz", "Iz", "Iy", "Jxx"],
        )
        properties[:, SectionAggregatorProperties.h:SectionAggregatorProperties.Jxx+1] = sec_props
        sec_aggregator = Sec_Aggregator(
            index = index,
            sec_name = sec_name,
            sec_shape = sec_shape,
            base_sec_class = base_sec_class,
            base_sec_idx = base_sec_idx,
            mats_class = mats_class,
            mats_idx = mats_idx,
            mat_type = mat_type,
            sec_model = sec_model,
            aggregated_sec_class = aggregated_sec_class,
            aggregated_sec_idx = aggregated_sec_idx,
            properties = properties,
            name_to_idx = name_to_idx,
        ) # Storing sections data to dataclass
        self._secs_list.append(sec_aggregator) # Append Sec_Aggregator Properties into Section Lists
        return sec_aggregator

    def _translate_slab_sections(self):
        sheet_name = "Slab Sections"
        data = self._reader.read(sheet_name=sheet_name, start_row=6) # Reading Sheet "Slab Sections" in the Input file
        if not self._validate_data(data=data, sheet_name=sheet_name, mandatory=False):
            slab_sections = []
            return slab_sections
        n = len(data)
        index = np.arange(n, dtype=np.int32)
        sec_name = np.empty(n, dtype="U32")
        mats_class = np.empty((n, 1), dtype=np.int32)
        mats_idx = np.empty((n, 1), dtype=np.int32)
        properties = np.zeros((n, len(SlabSectionProperties)), dtype=np.float64)
        name_to_idx = {}
        for i, row in enumerate(data):
            name = str(row["Section Name"])
            if name in name_to_idx:
                raise ValidationError(f"Duplicate Section name '{name}'")
            name_to_idx[name] = index[i]
            sec_name[i] = name
            mat_class, _, mat_idx = self._retrieve_material_index(mats_name=str(row["Material"]))
            mats_class[i] = mat_class
            mats_idx[i] = mat_idx
            t = self._to_internalunits.length(value=row["t"])
            properties[i] = (t)
        slab_sections = SlabSections(
            index = index,
            sec_name = sec_name,
            mats_class = mats_class,
            mats_idx = mats_idx,
            properties = properties,
            name_to_idx = name_to_idx,
        ) # Storing slab sections data to dataclass
        return slab_sections
    
    def _translate_point_objects(self, project_information):
        sheet_name = "Point Objects"
        data = self._reader.read(sheet_name=sheet_name, start_row=7) # Reading Sheet "Point Objects" in the Input file
        self._validate_data(data=data, sheet_name=sheet_name, mandatory=True)
        n = len(data)
        index = np.arange(n, dtype=np.int32)
        unique_name = np.empty(n, dtype="U15")
        coords = np.empty((n, 3), dtype=np.float64)
        name_to_idx = {}
        for i, row in enumerate(data):
            name = str(row["Unique Name"])
            if name in name_to_idx:
                raise ValidationError(f"Duplicate Point object name '{name}'")
            name_to_idx[name] = index[i]
            unique_name[i] = name
            coords[i] = (
                self._to_internalunits.length(value=row["X"]),
                self._to_internalunits.length(value=row["Y"]),
                self._to_internalunits.length(value=row["Z"]),
            )
        point_objects = {
            "Index": index,
            "Unique Name": unique_name,
            "Coordinates": coords,
            "Name to Index": name_to_idx,
        } # Storing point object data to dictionary
        ndim = project_information.ndim # Retrieve number of dimensional space
        storey_elevations = np.unique(coords[:, 2]) if ndim == 3 else np.unique(coords[:, 1]) # Determine storey elevations
        storeys = self._generate_storeys(storey_elevations=storey_elevations) # Generating Storey data from storey elevations
        return point_objects, storeys

    def _translate_line_objects(self, point_objects):
        sheet_name = "Line Objects"
        data = self._reader.read(sheet_name=sheet_name, start_row=15) # Reading Sheet "Line Objects" in the Input file
        self._validate_data(data=data, sheet_name=sheet_name, mandatory=True)
        n = len(data)
        index = np.arange(n, dtype=np.int32)
        unique_name = np.empty(n, dtype="U15")
        end_points_idx = np.empty((n, 2), dtype=np.int32)
        element_type = np.empty(n, dtype="U15")
        sec_class = np.empty(n, dtype=np.int32)
        sec_idx = np.empty(n, dtype=np.int32)
        is_auto_end_offsets = np.empty(n, dtype=bool)
        rigid_zone_factor = np.empty(n, dtype=np.float64)
        offsets_length = np.empty((n, 2), dtype=np.float64)
        is_zero_length_element = np.empty(n, dtype=bool)
        name_to_idx = {}
        for i, row in enumerate(data):
            name = str(row["Unique Name"])
            if name in name_to_idx:
                raise ValidationError(f"Duplicate Line object name '{name}'")
            name_to_idx[name] = index[i]
            unique_name[i] = name
            end_points_idx[i] = (point_objects["Name to Index"][str(row["I-End"])], point_objects["Name to Index"][str(row["J-End"])],)
            element_type[i] = (str(row["Element Type"]))
            is_auto_end_offsets[i] = (str(row["End Offset"]).strip().lower() == "auto from connectivity")
            rigid_zone_factor[i] = row["Rigid Zone Factor"]
            offsets_length[i] = (
                self._to_internalunits.length(value=row["I-End Offset Length"] if row["I-End Offset Length"] is not None else 0.0),
                self._to_internalunits.length(value=row["J-End Offset Length"] if row["J-End Offset Length"] is not None else 0.0),
            )
            sec_class[i], _, sec_idx[i] = self._retrieve_section_index(secs_name=str(row["Section"]))
            is_zero_length_element[i] = (str(row["Zero Length Element"]).strip().lower() == "yes")
        line_objects = {
            "Index": index,
            "Unique Name": unique_name,
            "End Points Index": end_points_idx,
            "Element Type": element_type,
            "Section Class": sec_class,
            "Section Index": sec_idx,
            "Is Auto End Offsets": is_auto_end_offsets,
            "Rigid Zone Factor": rigid_zone_factor,
            "Offsets Length": offsets_length,
            "Is Zero Length Element": is_zero_length_element,
            "Name to Index": name_to_idx,
        } # Storing line objects data to dictionary
        return line_objects

    def _translate_surface_objects(self, line_objects, slab_sections):
        sheet_name = "Surface Objects"
        data = self._reader.read(sheet_name=sheet_name, start_row=10) # Reading Sheet "Surface Objects" in the Input file
        self._validate_data(data=data, sheet_name=sheet_name, mandatory=True)
        n = len(data)
        index = np.arange(n, dtype=np.int32)
        unique_name = np.empty(n, dtype="U15")
        edges_idx = np.empty((n, 4), dtype=np.int32)
        vertices_idx = np.empty((n, 4), dtype=np.int32)
        element_type = np.empty(n, dtype="U15")
        sec_class = np.empty(n, dtype=np.int32)
        sec_idx = np.empty(n, dtype=np.int32)
        name_to_idx = {}
        for i, row in enumerate(data):
            name = str(row["Unique Name"])
            if name in name_to_idx:
                raise ValidationError(f"Duplicate Surface object name '{name}'")
            name_to_idx[name] = index[i]
            unique_name[i] = name
            edges_name = [edge for edge in (str(row["Edge 1"]), str(row["Edge 2"]), str(row["Edge 3"]), str(row["Edge 4"]),) if edge is not None]
            edges_idx[i], vertices_idx[i] = get_edges_and_vertices_from_surface(
                edges_name=edges_name,
                line_objects=line_objects,
                surface_name=name,
            )
            element_type[i] = (str(row["Element Type"]))
            sec_class[i], _, sec_idx[i] = self._retrieve_slabsection_index(
                secs_name=str(row["Section"]),
                secs_list=[slab_sections],
            )
        surface_objects = {
            "Index": index,
            "Unique Name": unique_name,
            "Edges Index": edges_idx,
            "Vertices Index": vertices_idx,
            "Element Type": element_type,
            "Section Class": sec_class,
            "Section Index": sec_idx,
            "Name to Index": name_to_idx,
        } # Storing surface objects data to dictionary
        return surface_objects
    
    def _translate_restraints(self, point_objects):
        sheet_name="Restraints"
        data = self._reader.read(sheet_name=sheet_name, start_row=10) # Reading Sheet "Restraints" in the Input file
        self._validate_data(data=data, sheet_name=sheet_name, mandatory=True)
        n = len(data)
        point_idx = np.empty(n, dtype=np.int32)
        dofs = np.empty((n, 6), dtype=np.int32)
        idx_set = set()
        for i, row in enumerate(data):
            point_name = str(row["Point"])
            idx = point_objects["Name to Index"][point_name]
            if idx in idx_set:
                raise ValidationError(f"Duplicate Restraints at Point '{point_name}'")
            idx_set.add(idx)
            point_idx[i] = idx
            dofs[i] = (
                int(row["UX"]),
                int(row["UY"]),
                int(row["UZ"]),
                int(row["RX"]),
                int(row["RY"]),
                int(row["RZ"]),
            )
        restraints = {
            "Point Index": point_idx,
            "DOFs": dofs,
        } # Storing restraints data to dictionary
        return restraints